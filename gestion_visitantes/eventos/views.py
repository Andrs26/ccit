import json, base64
from django.shortcuts import redirect, render
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
import openpyxl
from .models import EventoCapacitacion, EventoVisitante
from django.core.mail import send_mail
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
User = get_user_model()

# Create your views here.
@login_required
def guardar_evento(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        organizador = request.POST.get("organizador")

        organizador_id = request.POST.get("organizador")
        organizador_instance = User.objects.get(pk=organizador_id)

        fecha = request.POST.get("fecha")  # formato YYYY-MM-DD
        hora_inicio = request.POST.get("hora_inicio")
        hora_fin = request.POST.get("hora_fin")

        estado = 'activo'
        
        # Recopilar participantes ingresados manualmente (en formato JSON)
        participantes_json = request.POST.get("visitantes_data", "[]")
        try:
            participantes_manual = json.loads(participantes_json)
        except json.JSONDecodeError:
            participantes_manual = []
        
        # Procesar archivo Excel si se adjunta
        participantes_excel = []
        if "participantes_file" in request.FILES:
            file = request.FILES["participantes_file"]
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                # Asumir que la primera fila es header, y las columnas son: Nombre, Documento (en ese orden)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre_excel = row[0] if row[0] is not None else ""
                    documento_excel = row[1] if row[1] is not None else ""
                    participantes_excel.append({
                        "nombre": str(nombre_excel).strip(),
                        "documento": str(documento_excel).strip()
                    })
            except Exception as e:
                messages.error(request, "Error al procesar el archivo Excel: " + str(e))
                return redirect(request.META.get('HTTP_REFERER', '/'))
        
        # Combinar participantes manuales y del Excel
        todos_participantes = participantes_manual + participantes_excel
        
        # Eliminar duplicados: si hay un participante con documento, usarlo; si no, por nombre.
        participantes_unicos = []
        seen = set()
        for p in todos_participantes:
            key = p.get("documento") if p.get("documento") else p.get("nombre")
            if key not in seen:
                seen.add(key)
                participantes_unicos.append(p)
        
        # Crear el evento
        evento = EventoCapacitacion.objects.create(
            nombre = nombre,
            organizador = organizador_instance,
            fecha = fecha,
            hora_inicio = hora_inicio,
            hora_fin = hora_fin,
            estado = estado,
            cantidad_visitantes = len(participantes_unicos)
        )
        
        # Crear registros para cada participante
        for p in participantes_unicos:
            EventoVisitante.objects.create(
                id_evento = str(evento.id),  # o puedes usar otro identificador si tienes
                nombre_visitante = p.get("nombre", ""),
                documento_identificacion = p.get("documento", ""),
                cat_participante = ""
            )

        # Enviar correo al usuario en recepción
        try:
            # Buscar el primer usuario que pertenezca al grupo "visitas_recepcion_group"
            usuario_recepcion = User.objects.filter(groups__name='visitas_recepcion_group').first()
            
            if not usuario_recepcion:
                print("No se encontró ningún usuario en el grupo visitas_recepcion_group")
            else:
                usuario_destino = usuario_recepcion.email  # Correo del usuario destino
                
                # Obtener el usuario organizador a partir del username contenido en la variable "organizador"
                usuario_organizador = User.objects.get(id=organizador)
                
                subject = "Notificación: Nuevo Evento"
                message = (
                    f"Hola {usuario_recepcion.first_name},\n\n"
                    f"Se ha registrado el evento '{nombre}' con fecha: {fecha} en horario: {hora_inicio} - {hora_fin}, "
                    f"organizado por '{usuario_organizador.first_name} {usuario_organizador.last_name}'.\n\n"
                    "Saludos,\nEquipo CCIT"
                )
                send_mail(
                    subject,
                    message,
                    None,  # O puedes usar DEFAULT_FROM_EMAIL si lo configuraste en settings.py
                    [usuario_destino],
                    fail_silently=False,
                )
                print("Correo enviado a", usuario_destino)
        except User.DoesNotExist:
            print("Usuario no encontrado para enviar correo.")
        
        messages.success(request, "Evento registrado exitosamente con " + str(len(participantes_unicos)) + " participantes.")
        return redirect('colaboradores_home')
    else:
        # Si es GET, puedes cargar datos adicionales que necesites
        return render(request, 'visitantes/colaborador/nuevo_evento.html', {})

import json, base64
import openpyxl
from django.shortcuts import redirect, get_object_or_404, render
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from .models import EventoCapacitacion, EventoVisitante

@login_required
def save_editar_evento(request):
    if request.method == "POST":
        id = request.POST.get("id")
        evento = get_object_or_404(EventoCapacitacion, id=id)

        nombre = request.POST.get("nombre")
        from django.contrib.auth.models import User  # Importar el modelo User

        # Buscar el usuario correspondiente al ID
        organizador_id = request.POST.get("organizador")
        try:
            organizador = User.objects.get(id=organizador_id)  # Obtener la instancia de User
        except User.DoesNotExist:
            messages.error(request, "El organizador seleccionado no es válido.")
            return redirect(request.META.get('HTTP_REFERER', '/'))
        
        fecha = request.POST.get("fecha")  # formato YYYY-MM-DD
        hora_inicio = request.POST.get("hora_inicio")
        hora_fin = request.POST.get("hora_fin")
        
        # Recopilar participantes ingresados manualmente (en formato JSON)
        participantes_json = request.POST.get("visitantes_data", "[]")
        try:
            participantes_manual = json.loads(participantes_json)
        except json.JSONDecodeError:
            participantes_manual = []
        
        # Procesar archivo Excel si se adjunta
        participantes_excel = []
        if "participantes_file" in request.FILES:
            file = request.FILES["participantes_file"]
            try:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                # Asumir que la primera fila es header, y las columnas son: Nombre, Documento (en ese orden)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre_excel = row[0] if row[0] is not None else ""
                    documento_excel = row[1] if row[1] is not None else ""
                    participantes_excel.append({
                        "nombre": str(nombre_excel).strip(),
                        "documento": str(documento_excel).strip()
                    })
            except Exception as e:
                messages.error(request, "Error al procesar el archivo Excel: " + str(e))
                return redirect(request.META.get('HTTP_REFERER', '/'))
        
        # Combinar participantes manuales y del Excel
        todos_participantes = participantes_manual + participantes_excel
        
        # Eliminar duplicados: si hay un participante con documento, usarlo; si no, por nombre.
        participantes_unicos = []
        seen = set()
        for p in todos_participantes:
            key = p.get("documento") if p.get("documento") else p.get("nombre")
            if key not in seen:
                seen.add(key)
                participantes_unicos.append(p)
        
        # Editar el evento
        evento.nombre = nombre
        evento.organizador = organizador
        evento.fecha = fecha
        evento.hora_inicio = hora_inicio
        evento.hora_fin = hora_fin
        evento.cantidad_visitantes = len(participantes_unicos)
        evento.save()
        
        # Actualizar participantes:
        # Primero, eliminar los registros de EventoVisitante existentes para este evento
        EventoVisitante.objects.filter(id_evento=str(evento.id)).delete()
        
        # Luego, crear nuevos registros con los participantes únicos
        for p in participantes_unicos:
            EventoVisitante.objects.create(
                id_evento = str(evento.id),
                nombre_visitante = p.get("nombre", ""),
                documento_identificacion = p.get("documento", ""),
                cat_participante = ""
            )
        
        messages.success(request, "Evento modificado exitosamente.")
        return redirect('detalles_evento_visitas', id)
    else:
        messages.error(request, "Algo no va bien")
        return redirect(request.META.get('HTTP_REFERER', '/'))
    
@login_required
def editar_evento(request, id):
    evento = EventoCapacitacion.objects.get(id=id)
    participantes_evento = EventoVisitante.objects.filter(id_evento=id)

    return render(request, 'visitantes/recepcion/secundarios/editar_evento_colaborador.html', {
        'evento': evento,
        'participantes_evento': participantes_evento,
    })

@login_required
def cancelar_evento(request, id):
    evento = get_object_or_404(EventoCapacitacion, id=id)
    estado = 'cancelado'
    evento.estado = estado
    evento.save()
    
    messages.success(request, "Evento cancelado.")
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
def reanudar_evento(request, id):
    evento = get_object_or_404(EventoCapacitacion, id=id)
    if evento.estado == 'cancelado':
        estado = 'activo'
        evento.estado = estado
        evento.save()
    
        messages.success(request, "Evento cancelado.")
        return redirect(request.META.get('HTTP_REFERER', '/'))
    else:
        messages.error(request, "No se puede completar la acción.")
        return redirect(request.META.get('HTTP_REFERER', '/'))

import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import EventoVisitante

@require_POST
def actualizar_participante(request, participante_id):
    try:
        participante = EventoVisitante.objects.get(id=participante_id)
    except EventoVisitante.DoesNotExist:
        return JsonResponse({"success": False, "message": "Participante no encontrado."}, status=404)
    
    # Se reciben los campos "entrada" y/o "salida" en POST (como strings "true" o "false")
    entrada = request.POST.get('entrada')
    salida = request.POST.get('salida')
    
    # Obtener la hora actual (solo la parte de la hora)
    current_time = timezone.localtime().time()
    
    if entrada is not None:
        new_entrada = entrada.lower() == 'true'
        participante.entrada = new_entrada
        # Si se marca entrada, se asigna la hora actual; si se desmarca, se limpia el campo.
        if new_entrada:
            participante.hora_entrada = current_time
        else:
            participante.hora_entrada = None

    if salida is not None:
        new_salida = salida.lower() == 'true'
        # Solo permitimos marcar salida si la entrada ya está activa (esto ya se controla en el cliente)
        participante.salida = new_salida
        if new_salida:
            participante.hora_salida = current_time
        else:
            participante.hora_salida = None

    participante.save()
    return JsonResponse({
        "success": True,
        "estado": {
            "entrada": participante.entrada,
            "salida": participante.salida,
            "hora_entrada": participante.hora_entrada.strftime("%H:%M:%S") if participante.hora_entrada else "",
            "hora_salida": participante.hora_salida.strftime("%H:%M:%S") if participante.hora_salida else ""
        }
    })